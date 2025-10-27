# File to create a confusion matrix based on manual scoring
import pandas as pd
from openpyxl import load_workbook

# ----------------- File paths -----------------
input_file = r"D:\Vetology\Research Student Assignments\Research Student Assignments\Input Data 2 - feline_thorax_scoring50.xlsx"
output_file = r"D:\Vetology\Research Student Assignments\Research Student Assignments\confusion_output.xlsx"
# ----------------------------------------------

# Canonical condition list
conditions = [
    "pulmonary_nodules",
    "esophagitis",
    "pneumonia",
    "bronchitis",
    "interstitial",
    "diseased_lungs",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "perihilar_infiltrate",
    "rtm",
    "focal_caudodorsal_lung",
    "right_sided_cardiomegaly",
    "focal_perihilar",
    "left_sided_cardiomegaly",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "thoracic_lymphadenopathy",
    "pulmonary_hypoinflation",
    "pericardial_effusion",
    "Fe_Alveolar",
]

# Optional alias map for human-readable phrases
alias_map = {
    "bronchial pattern": "bronchitis",
    "interstitial pattern": "interstitial",
    "alveolar pattern": "Fe_Alveolar",
    "pleural fluid": "pleural_effusion",
    "pleural fissure lines": "pleural_effusion",
    "cardiac enlargement": "cardiomegaly",
    "cardiomegaly suspected": "cardiomegaly",
    "pulmonary vasculature enlargement": "pulmonary_vessel_enlargement",
}

# Read Excel
df = pd.read_excel(input_file)
print("✅ Columns detected:", list(df.columns))

# Map to your columns (update names if needed)
col_map = {"tp": "True_Pos", "tn": "True_Neg", "fp": "False_Pos", "fn": "False_Neg"}

# Fallback for missing or misnamed columns
for key, col in col_map.items():
    if col not in df.columns:
        print(f"⚠️ Column '{col}' not found in file. Please verify header names.")
        df[col] = ""


# Utility: clean & split a cell into tokens (no regex)
def split_clean(cell):
    if pd.isna(cell):
        return []
    s = str(cell).strip().lower()

    # Normalize separators — replace everything we can think of with commas
    for sep in [";", "|", "/", "\\", "\n", "\r", "&", " and "]:
        s = s.replace(sep, ",")
    # Replace double spaces and weird unicode spaces
    s = s.replace(" ", " ").replace("  ", " ")

    # Replace multiple commas with single
    while ",," in s:
        s = s.replace(",,", ",")

    # Split and trim
    tokens = [x.strip() for x in s.split(",") if x.strip()]

    # Apply alias normalization
    cleaned = []
    for t in tokens:
        if t in alias_map:
            cleaned.append(alias_map[t])
        else:
            cleaned.append(t)

    # Deduplicate while preserving order
    seen = set()
    uniq = []
    for x in cleaned:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


# Count initialization
counts = {cond: {"tp": 0, "fn": 0, "tn": 0, "fp": 0} for cond in conditions}

# Iterate through dataframe rows
for _, row in df.iterrows():
    # Get lists of tokens for each classification
    tp_tokens = split_clean(row[col_map["tp"]])
    tn_tokens = split_clean(row[col_map["tn"]])
    fp_tokens = split_clean(row[col_map["fp"]])
    fn_tokens = split_clean(row[col_map["fn"]])

    for cond in conditions:
        key = cond.lower()
        if key in tp_tokens:
            counts[cond]["tp"] += 1
        if key in tn_tokens:
            counts[cond]["tn"] += 1
        if key in fp_tokens:
            counts[cond]["fp"] += 1
        if key in fn_tokens:
            counts[cond]["fn"] += 1

# Convert to DataFrame
output_data = []
for cond, c in counts.items():
    output_data.append([cond, c["tp"], c["fn"], c["tn"], c["fp"], "", "", "", "", "", "", ""])

output_columns = [
    "Condition",
    "tp_Positive",
    "fn_Positive",
    "tn_Positive",
    "fp_Positive",
    "Sensitivity",
    "Specificity",
    "Check",
    "Positive Ground Truth",
    "Negative Ground Truth",
    "Ground Truth Check",
    "Radiologist Agreement Rate",
]

out_df = pd.DataFrame(output_data, columns=output_columns)
out_df.to_excel(output_file, index=False)

# Add formulas
wb = load_workbook(output_file)
ws = wb.active
for i in range(2, len(out_df) + 2):
    ws[f"F{i}"] = f"=IFERROR(B{i}/(B{i}+C{i}), 0)"
    ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+E{i}), 0)"
    ws[f"H{i}"] = f"=SUM(B{i}:E{i})"
    ws[f"I{i}"] = f"=SUM(B{i}:C{i})"
    ws[f"J{i}"] = f"=SUM(D{i}:E{i})"
    ws[f"K{i}"] = f"=SUM(I{i}:J{i})"
    ws[f"L{i}"] = 0  # Radiologist agreement rate = 0

wb.save(output_file)
print("✅ Confusion matrix created successfully at:", output_file)
