# File to use LLMs for scoring and creating the confusion matrix. Made calls to OpenAI and Gemini in this file. Will update once I try using Claude or open source models.

import pandas as pd
import json
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import os
import google.generativeai as genai

# Load environment variables
load_dotenv()

# Initialize OpenAI client
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

# Medical conditions to evaluate
CONDITIONS = [
    "pulmonary_nodules",
    "esophagitis",
    "pneumonia",
    "bronchitis",
    "interstitial",
    "diseased_lungs",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "perihilar_infiltrate",
    "rtm",
    "focal_caudodorsal_lung",
    "right_sided_cardiomegaly",
    "focal_perihilar",
    "left_sided_cardiomegaly",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "thoracic_lymphadenopathy",
    "pulmonary_hypoinflation",
    "pericardial_effusion",
    "Fe_Alveolar",
]

# Input prompt configuration and sheet settings for excel
input_prompt_file = "config.json"
sheet_name = "Sheet2"


def load_prompt_from_json(json_file_path=input_prompt_file):
    """
    Load system prompt from JSON file
    """
    try:
        with open(json_file_path, "r", encoding="utf-8") as f:
            prompts = json.load(f)
        return prompts.get("system_prompt", "")
    except FileNotFoundError:
        print(f"Error: {json_file_path} not found!")
        return None
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in {json_file_path}")
        return None


def process_radiology_data_openai(input_file_path, sheet_name=sheet_name):
    """
    Read radiology data from Excel/CSV file and process each case
    """
    # Read input file - specify sheet for Excel files
    if input_file_path.endswith(".csv"):
        df = pd.read_csv(input_file_path)
    else:
        # Read from specific sheet (Sheet2)
        df = pd.read_excel(input_file_path, sheet_name=sheet_name)

    print(f"Reading from sheet: {sheet_name}")
    print(f"Total rows: {len(df)}")

    # Ensure required columns exist
    required_columns = ["Findings (original radiologist report)", "Findings (AI report)"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Load system prompt from JSON
    system_prompt = load_prompt_from_json()
    if system_prompt is None:
        raise ValueError("Failed to load system prompt from JSON file")

    all_results = []

    # Process each case
    for idx, row in df.iterrows():
        radiologist_report = row["Findings (original radiologist report)"]
        ai_report = row["Findings (AI report)"]

        print(f"Processing row {idx + 1}/{len(df)}...")

        user_prompt = f"""
                        Analyze this radiology case:

                        Findings (original radiologist report):
                        {radiologist_report}

                        Findings (AI report):
                        {ai_report}

                        Classify each of the {len(CONDITIONS)} medical conditions according to the rules.
                        """

        try:
            response = client.chat.completions.create(
                model="gpt-4o",  # or "gpt-4-turbo" or "gpt-3.5-turbo"
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                temperature=0.1,
                response_format={"type": "json_object"},
            )

            result = json.loads(response.choices[0].message.content)
            all_results.append(result)

        except Exception as e:
            print(f"Error processing row {idx + 1}: {str(e)}")
            continue

    return all_results


def process_radiology_data_gemini(input_file_path, sheet_name=sheet_name):
    """
    Read radiology data from Excel/CSV file and process each case using Gemini API
    """
    # Read input file
    if input_file_path.endswith(".csv"):
        df = pd.read_csv(input_file_path)
    else:
        df = pd.read_excel(input_file_path, sheet_name=sheet_name)

    print(f"Reading from sheet: {sheet_name}")
    print(f"Total rows: {len(df)}")

    required_columns = ["Findings (original radiologist report)", "Findings (AI report)"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Load system prompt
    system_prompt = load_prompt_from_json()
    if system_prompt is None:
        raise ValueError("Failed to load system prompt from JSON file")

    all_results = []

    # Create Gemini model
    model = genai.GenerativeModel("gemini-2.5-pro")

    for idx, row in df.iterrows():
        radiologist_report = row["Findings (original radiologist report)"]
        ai_report = row["Findings (AI report)"]

        print(f"Processing row {idx + 1}/{len(df)} (Gemini)...")

        user_prompt = f"""
                        Analyze this radiology case:

                        Findings (original radiologist report):
                        {radiologist_report}

                        Findings (AI report):
                        {ai_report}

                        Classify each of the {len(CONDITIONS)} medical conditions according to the rules.
                        The output MUST be a JSON object strictly in this structure:
                        {{
                          "conditions": {{
                            "condition_name": {{
                              "classification": "TP" | "FP" | "TN" | "FN"
                            }},
                            ...
                          }}
                        }}
                        """

        try:
            response = model.generate_content(
                contents=[{"role": "system", "parts": [system_prompt]}, {"role": "user", "parts": [user_prompt]}],
                generation_config=genai.GenerationConfig(temperature=0.1, response_mime_type="application/json"),
            )

            result = json.loads(response.text)
            all_results.append(result)

        except Exception as e:
            print(f"Error processing row {idx + 1}: {str(e)}")
            continue

    return all_results


def calculate_confusion_matrix(results):
    """
    Calculate cumulative TP, FP, TN, FN, Sensitivity, and Specificity for each condition
    """
    condition_stats = {condition: {"TP": 0, "FP": 0, "TN": 0, "FN": 0} for condition in CONDITIONS}

    # Aggregate results across all cases
    for case_result in results:
        conditions = case_result.get("conditions", {})
        for condition, data in conditions.items():
            if condition in condition_stats:
                classification = data.get("classification", "TN")
                condition_stats[condition][classification] += 1

    # Calculate metrics
    matrix_data = []
    for condition in CONDITIONS:
        stats = condition_stats[condition]
        tp = stats["TP"]
        fp = stats["FP"]
        tn = stats["TN"]
        fn = stats["FN"]

        # Calculate Sensitivity and Specificity
        sensitivity = (tp / (tp + fn) * 100) if (tp + fn) > 0 else 0
        specificity = (tn / (tn + fp) * 100) if (tn + fp) > 0 else 0

        # Calculate checks
        check = tp + fp + tn + fn
        positive_ground_truth = tp + fn
        negative_ground_truth = tn + fp
        ground_truth_check = positive_ground_truth + negative_ground_truth

        # Calculate radiologist agreement (assuming TP + TN are agreements)
        radiologist_agreement = ((tp + tn) / check * 100) if check > 0 else 0

        matrix_data.append(
            {
                "condition": condition,
                "tp_Positive": tp,
                "fn_Positive": fn,
                "tn_Positive": tn,
                "fp_Positive": fp,
                "Sensitivity": round(sensitivity, 2),
                "Specificity": round(specificity, 2),
                "Check": check,
                "Positive Ground Truth": positive_ground_truth,
                "Negative Ground Truth": negative_ground_truth,
                "Ground Truth Check": ground_truth_check,
                "Radiologist Agreement": f"{round(radiologist_agreement, 0)}%",
            }
        )

    return pd.DataFrame(matrix_data)


def export_to_excel(df, output_path="confusion_matrix_output.xlsx"):
    """
    Export confusion matrix to Excel with formatting
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Confusion Matrix"

    # Define header style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    column_widths = {
        "A": 25,  # condition
        "B": 12,  # tp_Positive
        "C": 12,  # fn_Positive
        "D": 12,  # tn_Positive
        "E": 12,  # fp_Positive
        "F": 12,  # Sensitivity
        "G": 12,  # Specificity
        "H": 10,  # Check
        "I": 18,  # Positive Ground Truth
        "J": 18,  # Negative Ground Truth
        "K": 18,  # Ground Truth Check
        "L": 18,  # Radiologist Agreement
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save workbook
    wb.save(output_path)
    print(f"Confusion matrix exported to {output_path}")


def main(use_gemini=False):
    """
    Main execution function
    """
    # Input file path
    input_file = "D:\Vetology\Research Student Assignments\Research Student Assignments\Input Data 2 - feline_thorax_scoring50.xlsx"
    output_file = "confusion_matrix_output_ai.xlsx"
    sheet_name = "Sheet2"  # Specify which sheet to read from

    print("Starting radiology data processing...")

    if use_gemini:
        results = process_radiology_data_gemini(input_file, sheet_name=sheet_name)

    else:
        results = process_radiology_data_openai(input_file, sheet_name=sheet_name)

    # Process all cases

    print(f"\nProcessed {len(results)} cases")

    # Calculate confusion matrix
    print("\nCalculating confusion matrix...")
    confusion_matrix_df = calculate_confusion_matrix(results)

    # Display summary
    print("\nConfusion Matrix Summary:")
    print(confusion_matrix_df.head())

    # Export to Excel
    export_to_excel(confusion_matrix_df, output_file)

    print("\nDone!")


if __name__ == "__main__":
    main(use_gemini=False)
