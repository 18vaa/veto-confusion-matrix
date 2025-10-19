"""
utils.py

Utility functions for reading scoring Excel files, counting condition
occurrences (TP/FN/TN/FP), building the output DataFrame, and writing
the result to an Excel file with formulas.

Drop this file in the same folder as your scripts and import the functions.
"""

from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook


def read_input_excel(path: str, nrows: Optional[int] = None, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Read an Excel file into a DataFrame.
    - path: full path to the Excel file
    - nrows: optional number of rows to read (useful for testing)
    - sheet_name: optional sheet name (default: first sheet)
    """
    df = pd.read_excel(path, sheet_name=sheet_name)
    if nrows is not None:
        df = df.head(nrows)
    return df


def count_conditions(
    df: pd.DataFrame, conditions: List[str], col_names: Dict[str, str] = None
) -> Dict[str, Dict[str, int]]:
    """
    Count occurrences of each condition in TP, FN, TN, FP columns.

    - df: input DataFrame
    - conditions: list of condition keys (strings) to search for
    - col_names: mapping for column names:
        {
            "tp": "True Positive",
            "fn": "False Negative",
            "tn": "True Negative",
            "fp": "False Positive"
        }
      If not provided, defaults to the above.
    Returns:
      counts dict: { condition: {"tp": int, "fn": int, "tn": int, "fp": int}, ... }
    """
    if col_names is None:
        col_names = {
            "tp": "True Positive",
            "fn": "False Negative",
            "tn": "True Negative",
            "fp": "False Positive",
        }

    # ensure keys exist (if missing, .get will return "")
    counts = {condition: {"tp": 0, "fn": 0, "tn": 0, "fp": 0} for condition in conditions}

    # iterate rows
    for _, row in df.iterrows():
        # get cell text or empty string, convert to lowercase for case-insensitive matching
        fp_cell = str(row.get(col_names["fp"], "")).lower()
        fn_cell = str(row.get(col_names["fn"], "")).lower()
        tp_cell = str(row.get(col_names["tp"], "")).lower()
        tn_cell = str(row.get(col_names["tn"], "")).lower()

        for cond in conditions:
            # match raw condition string presence in each cell
            # you can extend this to regex or exact matching if needed
            if cond.lower() in fp_cell:
                counts[cond]["fp"] += 1
            if cond.lower() in fn_cell:
                counts[cond]["fn"] += 1
            if cond.lower() in tp_cell:
                counts[cond]["tp"] += 1
            if cond.lower() in tn_cell:
                counts[cond]["tn"] += 1

    return counts


def counts_to_dataframe(counts: Dict[str, Dict[str, int]]) -> pd.DataFrame:
    """
    Convert the counts dictionary into a DataFrame with placeholder columns
    for formulas to be written later in Excel.
    """
    output_data = []
    for cond, c in counts.items():
        output_data.append(
            [
                cond,
                c["tp"],  # True Positive -> B
                c["fn"],  # False Negative -> C
                c["tn"],  # True Negative -> D
                c["fp"],  # False Positive -> E
                "",  # Sensitivity -> F
                "",  # Specificity -> G
                "",  # Check -> H
                "",  # Positive Ground Truth -> I
                "",  # Negative Ground Truth -> J
                "",  # Ground Truth Check -> K
                "",  # Radiologist Agreement Rate -> L
            ]
        )

    output_columns = [
        "Condition",
        "tp_Positive",
        "fn_Positive",
        "tn_Positive",
        "fp_Positive",
        "Sensitivity",
        "Specificity",
        "Check",
        "Positive Ground Truth",
        "Negative Ground Truth",
        "Ground Truth Check",
        "Radiologist Agreement Rate",
    ]

    return pd.DataFrame(output_data, columns=output_columns)


def write_df_with_formulas(output_df: pd.DataFrame, output_path: str) -> None:
    """
    Write the DataFrame to Excel and populate formula columns using openpyxl.

    The function writes the DataFrame first (so numeric counts are in place),
    then opens the workbook and writes formulas into columns F..L for each row.
    """
    # Save dataframe first (this creates header + values)
    output_df.to_excel(output_path, index=False)

    # Load workbook and write formulas
    wb = load_workbook(output_path)
    ws = wb.active

    n_rows = len(output_df)
    # Excel rows start at 1, header is row 1, data starts at row 2
    for i in range(2, n_rows + 2):
        ws[f"F{i}"] = f"=IFERROR(B{i}/(B{i}+C{i}), 0)"  # Sensitivity
        ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+E{i}), 0)"  # Specificity
        ws[f"H{i}"] = f"=SUM(B{i}:E{i})"  # Check (TP+FN+TN+FP)
        ws[f"I{i}"] = f"=SUM(B{i}:C{i})"  # Positive Ground Truth (TP+FN)
        ws[f"J{i}"] = f"=SUM(D{i}:E{i})"  # Negative Ground Truth (TN+FP)
        ws[f"K{i}"] = f"=SUM(I{i}:J{i})"  # Ground Truth Check
        ws[f"L{i}"] = f"=IFERROR((B{i}+D{i})/H{i}, 0)"  # Radiologist Agreement Rate

    wb.save(output_path)


def build_and_write_confusion(
    input_path: str,
    output_path: str,
    conditions: List[str],
    col_names: Dict[str, str] = None,
    nrows: Optional[int] = None,
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:
    """
    End-to-end helper: read input, count conditions, build DataFrame, and write output Excel with formulas.
    Returns the output DataFrame (without formulas evaluated).
    """
    df = read_input_excel(input_path, nrows=nrows, sheet_name=sheet_name)
    counts = count_conditions(df, conditions, col_names=col_names)
    output_df = counts_to_dataframe(counts)
    write_df_with_formulas(output_df, output_path)
    return output_df
